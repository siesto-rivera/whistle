from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import formats
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView

from .forms import WhistleCaseForm, WhistleTimelineForm, WhistleArticleForm, WhistleCheerForm
from .models import WhistleCase, WhistleTimeline, WhistleArticle, WhistleCheer


# ── 공개 페이지 (로그인 불필요) ────────────────────────────────


class AboutView(TemplateView):
    template_name = "whistle/about.html"


class WhistleHomeView(TemplateView):
    template_name = "whistle/home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["recent_cases"] = WhistleCase.objects.filter(hide=False).order_by("-case_year", "-id")[:6]
        context["recent_cheers"] = WhistleCheer.objects.select_related("case").order_by("-created_at")[:6]
        context["total_cases"] = WhistleCase.objects.filter(hide=False).count()
        context["total_cheers"] = WhistleCheer.objects.count()
        return context


class PublicWhistleListView(ListView):
    model = WhistleCase
    context_object_name = "cases"
    template_name = "whistle/public_list.html"
    paginate_by = 12

    def get_queryset(self):
        queryset = WhistleCase.objects.filter(hide=False).order_by("-case_year", "-id")
        q = self.request.GET.get("q", "").strip()
        category = self.request.GET.get("category", "").strip()
        organization = self.request.GET.get("organization", "").strip()
        if q:
            queryset = queryset.filter(
                Q(title__icontains=q)
                | Q(whistleblower__icontains=q)
                | Q(organization__icontains=q)
                | Q(category__icontains=q)
                | Q(tags__icontains=q)
                | Q(content__icontains=q)
                | Q(situation__icontains=q)
                | Q(awards__icontains=q)
                | Q(support__icontains=q)
                | Q(quote__icontains=q)
                | Q(media_coverage__icontains=q)
                | Q(media_detail__icontains=q)
            )
        if category:
            queryset = queryset.filter(category=category)
        if organization:
            queryset = queryset.filter(organization=organization)
        tag = self.request.GET.get("tag", "").strip()
        if tag:
            queryset = queryset.filter(tags__icontains=tag)
        return queryset

    def _all_tags(self):
        tags = set()
        for val in WhistleCase.objects.filter(hide=False).values_list("tags", flat=True):
            if not val:
                continue
            for t in val.split(","):
                t = t.strip()
                if t:
                    tags.add(t)
        return sorted(tags)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "").strip()
        context["current_category"] = self.request.GET.get("category", "").strip()
        context["current_organization"] = self.request.GET.get("organization", "").strip()
        context["current_tag"] = self.request.GET.get("tag", "").strip()
        context["categories"] = WhistleCase.CATEGORY_CHOICES
        context["organizations"] = WhistleCase.ORGANIZATION_CHOICES
        context["all_tags"] = self._all_tags()
        return context


class PublicWhistleDetailView(DetailView):
    model = WhistleCase
    template_name = "whistle/public_detail.html"
    context_object_name = "case"

    def get_queryset(self):
        return WhistleCase.objects.filter(hide=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["timelines"] = WhistleTimeline.objects.filter(case=self.object).order_by("-rdate")
        context["articles"] = WhistleArticle.objects.filter(case=self.object)
        context["cheer_form"] = WhistleCheerForm()
        if self.object.tags:
            context["tag_list"] = [t.strip() for t in self.object.tags.split(",") if t.strip()]
        return context


# ── API ──────────────────────────────────────────────────────


def whistle_search_api(request):
    """공익제보 사건 제목 검색 JSON API — 자동완성용"""
    q = request.GET.get("q", "").strip()
    if len(q) < 1:
        return JsonResponse([], safe=False)
    cases = WhistleCase.objects.filter(title__icontains=q).order_by("-case_year", "-id")[:20]
    results = [
        {"id": c.pk, "title": c.title}
        for c in cases
    ]
    return JsonResponse(results, safe=False)


@login_required
def tag_list_api(request):
    """기존 태그 목록 JSON API"""
    tags = set()
    for val in WhistleCase.objects.values_list("tags", flat=True):
        if not val:
            continue
        for t in val.split(","):
            t = t.strip()
            if t:
                tags.add(t)
    return JsonResponse(sorted(tags), safe=False)


# ── 관리 대시보드 ────────────────────────────────────────────────


class WhistleDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "whistle/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total_cases"] = WhistleCase.objects.count()
        context["public_cases"] = WhistleCase.objects.filter(hide=False).count()
        context["hidden_cases"] = WhistleCase.objects.filter(hide=True).count()
        context["total_timelines"] = WhistleTimeline.objects.count()
        context["total_articles"] = WhistleArticle.objects.count()
        context["total_cheers"] = WhistleCheer.objects.count()
        context["recent_cases"] = WhistleCase.objects.order_by("-id")[:5]
        context["recent_cheers"] = WhistleCheer.objects.select_related("case").order_by("-created_at")[:5]
        context["categories"] = (
            WhistleCase.objects.filter(hide=False)
            .values_list("category", flat=True)
        )
        # 카테고리별 건수
        from django.db.models import Count
        from collections import Counter
        context["category_stats"] = (
            WhistleCase.objects.filter(hide=False)
            .values("category")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        # 불이익상세 통계
        from django.utils.html import strip_tags
        disadvantage_counter = Counter()
        for val in WhistleCase.objects.exclude(hidden_disadvantage="").values_list("hidden_disadvantage", flat=True):
            text = strip_tags(val).strip()
            if text:
                for line in text.replace("<br>", "\n").split("\n"):
                    line = line.strip().strip("-").strip()
                    if line:
                        disadvantage_counter[line] += 1
        context["disadvantage_stats"] = disadvantage_counter.most_common()
        # 위반행위 통계
        violation_counter = Counter()
        for val in WhistleCase.objects.exclude(hidden_violation="").values_list("hidden_violation", flat=True):
            text = strip_tags(val).strip()
            if text:
                for line in text.replace("<br>", "\n").split("\n"):
                    line = line.strip().strip("-").strip()
                    if line:
                        violation_counter[line] += 1
        context["violation_stats"] = violation_counter.most_common()
        # 공익제보자상 수상 통계
        context["prize_total"] = WhistleCase.objects.filter(prize=True).count()
        context["prize_cases"] = WhistleCase.objects.filter(prize=True).order_by("-case_year", "-id")
        return context


# ── 불이익상세 통계 ────────────────────────────────────────────────


class DisadvantageStatsView(LoginRequiredMixin, TemplateView):
    template_name = "whistle/disadvantage_stats.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from collections import Counter
        from django.utils.html import strip_tags
        counter = Counter()
        case_map = {}
        for case in WhistleCase.objects.exclude(hidden_disadvantage=""):
            text = strip_tags(case.hidden_disadvantage).strip()
            if not text:
                continue
            for line in text.replace("<br>", "\n").split("\n"):
                line = line.strip().strip("-").strip()
                if line:
                    counter[line] += 1
                    case_map.setdefault(line, []).append(case)
        selected = self.request.GET.get("item", "").strip()
        context["stats"] = counter.most_common()
        context["total_items"] = len(counter)
        context["selected_item"] = selected
        if selected and selected in case_map:
            context["filtered_cases"] = case_map[selected]
        return context


# ── 위반행위 통계 ─────────────────────────────────────────────────


class ViolationStatsView(LoginRequiredMixin, TemplateView):
    template_name = "whistle/violation_stats.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from collections import Counter
        from django.utils.html import strip_tags
        counter = Counter()
        case_map = {}
        for case in WhistleCase.objects.exclude(hidden_violation=""):
            text = strip_tags(case.hidden_violation).strip()
            if not text:
                continue
            for line in text.replace("<br>", "\n").split("\n"):
                line = line.strip().strip("-").strip()
                if line:
                    counter[line] += 1
                    case_map.setdefault(line, []).append(case)
        selected = self.request.GET.get("item", "").strip()
        context["stats"] = counter.most_common()
        context["total_items"] = len(counter)
        context["selected_item"] = selected
        if selected and selected in case_map:
            context["filtered_cases"] = case_map[selected]
        return context


# ── 공익제보자상 통계 ─────────────────────────────────────────────


class PrizeStatsView(LoginRequiredMixin, TemplateView):
    template_name = "whistle/prize_stats.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["prize_cases"] = WhistleCase.objects.filter(prize=True).order_by("-case_year", "-id")
        context["prize_total"] = context["prize_cases"].count()
        context["total_cases"] = WhistleCase.objects.count()
        return context


# ── 태그 통계 ──────────────────────────────────────────────────


class TagStatsView(LoginRequiredMixin, TemplateView):
    template_name = "whistle/tag_stats.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from collections import Counter
        tag_counter = Counter()
        tag_cases = {}
        for case in WhistleCase.objects.all():
            if not case.tags:
                continue
            for tag in case.tags.split(","):
                tag = tag.strip()
                if not tag:
                    continue
                tag_counter[tag] += 1
                tag_cases.setdefault(tag, []).append(case)

        selected_tag = self.request.GET.get("tag", "").strip()
        sort = self.request.GET.get("sort", "name")
        if sort == "count":
            context["tag_stats"] = sorted(tag_counter.items(), key=lambda x: (-x[1], x[0]))
        else:
            context["tag_stats"] = sorted(tag_counter.items(), key=lambda x: x[0])
        context["total_tags"] = len(tag_counter)
        context["selected_tag"] = selected_tag
        context["current_sort"] = sort
        if selected_tag and selected_tag in tag_cases:
            context["filtered_cases"] = tag_cases[selected_tag]
        return context


# ── 엑셀 다운로드 ──────────────────────────────────────────────────


@login_required
def case_excel_download(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.utils.html import strip_tags

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공익제보 사건"

    headers = [
        "ID", "사건명", "제보년도", "제보자 성명", "신고대상",
        "공익침해분야", "주요태그", "제보내용", "제보자 상황",
        "수상이력", "참여연대 지원", "미디어_언론보도", "언론보도 내역",
        "미디어_사진", "제보자 한마디", "위반행위", "불이익상세",
        "참고", "비공개",
    ]
    ws.append(headers)

    # 헤더 스타일
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)

    for case in WhistleCase.objects.order_by("-case_year", "-id"):
        ws.append([
            case.id,
            case.title,
            case.case_year,
            case.whistleblower,
            case.organization,
            case.category,
            case.tags,
            strip_tags(case.content),
            strip_tags(case.situation),
            strip_tags(case.awards),
            strip_tags(case.support),
            strip_tags(case.media_coverage),
            strip_tags(case.media_detail),
            strip_tags(case.media_photo),
            strip_tags(case.quote),
            strip_tags(case.hidden_violation),
            strip_tags(case.hidden_disadvantage),
            strip_tags(case.memo),
            "Y" if case.hide else "N",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="whistle_cases.xlsx"'
    wb.save(response)
    return response


# ── 공익제보 사건 ─────────────────────────────────────────────────


class WhistleCaseListView(LoginRequiredMixin, ListView):
    model = WhistleCase
    context_object_name = "cases"
    template_name = "whistle/whistle_list.html"
    paginate_by = 30

    SORT_OPTIONS = {
        "title_asc": "title",
        "title_desc": "-title",
        "year_asc": "case_year",
        "year_desc": "-case_year",
        "hide_asc": "hide",
        "hide_desc": "-hide",
    }

    def get_queryset(self):
        sort = self.request.GET.get("sort", "year_desc")
        order = self.SORT_OPTIONS.get(sort, "-case_year")
        queryset = WhistleCase.objects.order_by(order, "-id")
        q = self.request.GET.get("q", "").strip()
        category = self.request.GET.get("category", "").strip()
        organization = self.request.GET.get("organization", "").strip()
        if q:
            queryset = queryset.filter(
                Q(title__icontains=q) | Q(whistleblower__icontains=q) | Q(tags__icontains=q)
            )
        if category:
            queryset = queryset.filter(category=category)
        if organization:
            queryset = queryset.filter(organization=organization)
        tag = self.request.GET.get("tag", "").strip()
        if tag:
            queryset = queryset.filter(tags__icontains=tag)
        return queryset

    def _all_tags(self):
        tags = set()
        for val in WhistleCase.objects.values_list("tags", flat=True):
            if not val:
                continue
            for t in val.split(","):
                t = t.strip()
                if t:
                    tags.add(t)
        return sorted(tags)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "").strip()
        context["current_category"] = self.request.GET.get("category", "").strip()
        context["current_organization"] = self.request.GET.get("organization", "").strip()
        context["current_tag"] = self.request.GET.get("tag", "").strip()
        context["current_sort"] = self.request.GET.get("sort", "year_desc")
        context["categories"] = WhistleCase.CATEGORY_CHOICES
        context["organizations"] = WhistleCase.ORGANIZATION_CHOICES
        context["all_tags"] = self._all_tags()
        return context


class WhistleCaseDetailView(LoginRequiredMixin, DetailView):
    model = WhistleCase
    template_name = "whistle/whistle_detail.html"
    context_object_name = "case"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["timelines"] = WhistleTimeline.objects.filter(case=self.object).order_by("-rdate")
        context["articles"] = WhistleArticle.objects.filter(case=self.object)
        context["timeline_form"] = WhistleTimelineForm(initial={"case": self.object.pk})
        context["article_form"] = WhistleArticleForm(initial={"case": self.object.pk})
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_type = request.POST.get("form_type")
        # 삭제
        if form_type == "delete_timeline":
            obj = get_object_or_404(WhistleTimeline, pk=request.POST.get("item_pk"), case=self.object)
            obj.delete()
            return redirect("whistle:case_detail", pk=self.object.pk)
        if form_type == "delete_article":
            obj = get_object_or_404(WhistleArticle, pk=request.POST.get("item_pk"), case=self.object)
            obj.delete()
            return redirect("whistle:case_detail", pk=self.object.pk)
        # 수정
        if form_type == "edit_timeline":
            obj = get_object_or_404(WhistleTimeline, pk=request.POST.get("item_pk"), case=self.object)
            form = WhistleTimelineForm(request.POST, instance=obj)
            if form.is_valid():
                form.save()
            return redirect("whistle:case_detail", pk=self.object.pk)
        if form_type == "edit_article":
            obj = get_object_or_404(WhistleArticle, pk=request.POST.get("item_pk"), case=self.object)
            form = WhistleArticleForm(request.POST, instance=obj)
            if form.is_valid():
                form.save()
            return redirect("whistle:case_detail", pk=self.object.pk)
        # 추가
        if form_type == "article":
            form = WhistleArticleForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect("whistle:case_detail", pk=self.object.pk)
            context = self.get_context_data()
            context["article_form"] = form
            return self.render_to_response(context)
        else:
            form = WhistleTimelineForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect("whistle:case_detail", pk=self.object.pk)
            context = self.get_context_data()
            context["timeline_form"] = form
            return self.render_to_response(context)


class WhistleCaseCreateView(LoginRequiredMixin, CreateView):
    model = WhistleCase
    form_class = WhistleCaseForm
    template_name = "whistle/whistle_form.html"

    def get_success_url(self):
        return reverse_lazy("whistle:case_detail", kwargs={"pk": self.object.pk})


class WhistleCaseUpdateView(LoginRequiredMixin, UpdateView):
    model = WhistleCase
    form_class = WhistleCaseForm
    template_name = "whistle/whistle_form.html"

    def get_success_url(self):
        return reverse_lazy("whistle:case_detail", kwargs={"pk": self.object.pk})


@login_required
@require_POST
def whistle_case_delete(request, pk):
    case = get_object_or_404(WhistleCase, pk=pk)
    case.delete()
    return redirect("whistle:case_list")


# ── 타임라인 ─────────────────────────────────────────────────


class WhistleTimelineListView(LoginRequiredMixin, ListView):
    model = WhistleTimeline
    context_object_name = "timelines"
    template_name = "whistle/timeline_list.html"
    paginate_by = 30

    def get_queryset(self):
        queryset = WhistleTimeline.objects.select_related("case").order_by("-rdate", "-id")
        q = self.request.GET.get("q", "").strip()
        if q:
            queryset = queryset.filter(
                Q(title__icontains=q) | Q(case__title__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "").strip()
        return context


class WhistleTimelineDetailView(LoginRequiredMixin, DetailView):
    model = WhistleTimeline
    template_name = "whistle/timeline_detail.html"
    context_object_name = "timeline"


class WhistleTimelineCreateView(LoginRequiredMixin, CreateView):
    model = WhistleTimeline
    form_class = WhistleTimelineForm
    template_name = "whistle/timeline_form.html"

    def get_success_url(self):
        return reverse_lazy("whistle:timeline_detail", kwargs={"pk": self.object.pk})


class WhistleTimelineUpdateView(LoginRequiredMixin, UpdateView):
    model = WhistleTimeline
    form_class = WhistleTimelineForm
    template_name = "whistle/timeline_form.html"

    def get_success_url(self):
        return reverse_lazy("whistle:timeline_detail", kwargs={"pk": self.object.pk})


# ── 관련기사 ─────────────────────────────────────────────────


class WhistleArticleListView(LoginRequiredMixin, ListView):
    model = WhistleArticle
    context_object_name = "articles"
    template_name = "whistle/article_list.html"
    paginate_by = 30

    def get_queryset(self):
        queryset = WhistleArticle.objects.select_related("case").order_by("-id")
        q = self.request.GET.get("q", "").strip()
        if q:
            queryset = queryset.filter(
                Q(title__icontains=q) | Q(case__title__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "").strip()
        return context


class WhistleArticleDetailView(LoginRequiredMixin, DetailView):
    model = WhistleArticle
    template_name = "whistle/article_detail.html"
    context_object_name = "article"


class WhistleArticleCreateView(LoginRequiredMixin, CreateView):
    model = WhistleArticle
    form_class = WhistleArticleForm
    template_name = "whistle/article_form.html"

    def get_success_url(self):
        return reverse_lazy("whistle:article_detail", kwargs={"pk": self.object.pk})


class WhistleArticleUpdateView(LoginRequiredMixin, UpdateView):
    model = WhistleArticle
    form_class = WhistleArticleForm
    template_name = "whistle/article_form.html"

    def get_success_url(self):
        return reverse_lazy("whistle:article_detail", kwargs={"pk": self.object.pk})


# ── 응원글 (공개) ────────────────────────────────────────────


def _format_dt(dt):
    return f"{dt.year}.{dt.month}.{dt.day} {dt.strftime('%H:%M')}"


def whistle_cheer_list_api(request, pk):
    """응원글 목록 JSON API"""
    case = get_object_or_404(WhistleCase, pk=pk, hide=False)
    cheers = case.cheers.all()[:50]
    data = [
        {
            "author_name": c.author_name,
            "content": c.content,
            "created_at": _format_dt(c.created_at),
        }
        for c in cheers
    ]
    return JsonResponse({"cheers": data, "count": len(data)})


@require_POST
def whistle_cheer_create(request, pk):
    case = get_object_or_404(WhistleCase, pk=pk, hide=False)
    form = WhistleCheerForm(request.POST)
    if form.is_valid():
        cheer = form.save(commit=False)
        cheer.case = case
        cheer.save()
        return JsonResponse({
            "ok": True,
            "cheer": {
                "author_name": cheer.author_name,
                "content": cheer.content,
                "created_at": _format_dt(cheer.created_at),
            },
        })
    return JsonResponse({"ok": False, "errors": form.errors}, status=400)


# ── 응원글 (관리자) ──────────────────────────────────────────


class WhistleCheerListView(LoginRequiredMixin, ListView):
    model = WhistleCheer
    context_object_name = "cheers"
    template_name = "whistle/cheer_list.html"
    paginate_by = 30

    def get_queryset(self):
        queryset = WhistleCheer.objects.select_related("case").order_by("-created_at")
        q = self.request.GET.get("q", "").strip()
        if q:
            queryset = queryset.filter(
                Q(author_name__icontains=q) | Q(case__title__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "").strip()
        return context


@login_required
@require_POST
def whistle_cheer_delete(request, pk):
    cheer = get_object_or_404(WhistleCheer, pk=pk)
    cheer.delete()
    return redirect("whistle:cheer_list")
